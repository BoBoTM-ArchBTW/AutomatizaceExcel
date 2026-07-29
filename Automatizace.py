import json
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import pandas as pd

# Soubor pro trvalé ukládání profilů
PROFILY_FILE = "konfigurace_profilu.json"

# Globální struktury
seznam_souboru_ui = []   # UI karty souborů na 1. stránce
nactene_soubory = {}     # Načtená data souborů {alias: {"cesta", "sheet", "hdr", "sloupce"}}
seznam_operaci = []      # Seznam kroků automatizace
cesta_cil_global = ""    # Výstupní cesta při práci s více soubory


# ================= HELPERY PRO PROFILY (JSON) =================
def nacist_profily():
    if not os.path.exists(PROFILY_FILE):
        return {}
    try:
        with open(PROFILY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def ulozit_profily(data):
    try:
        with open(PROFILY_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        messagebox.showerror("Chyba", f"Nepodařilo se uložit profily:\n{str(e)}")


# ================= POMOCNÉ UI A EXCEL HELPERY =================
def nastavit_scrolovani(canvas):
    """Připojí scrollování kolečkem myši k danému canvasu."""
    def _on_mousewheel(e):
        canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

    canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
    canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))


def pridat_napovedu(parent_frame, text):
    """Vytvoří zelený box s nápovědou nahoře v pop-up okně."""
    box = ttk.LabelFrame(parent_frame, text=" 💡 Nápověda ", padding=8)
    box.pack(fill=tk.X, pady=(0, 10))
    ttk.Label(
        box, text=text, font=("Segoe UI", 8, "italic"),
        wraplength=440, justify="left", foreground="#333333"
    ).pack(anchor=tk.W)


def nacist_excel(cesta, **kwargs):
    for eng in ["openpyxl", "xlrd", "pyxlsb", "calamine", None]:
        try:
            return pd.read_excel(cesta, engine=eng, **kwargs)
        except Exception:
            continue
    return None


def nacist_listy(cesta):
    for eng in ["openpyxl", "xlrd", "pyxlsb", "calamine", None]:
        try:
            return pd.ExcelFile(cesta, engine=eng).sheet_names
        except Exception:
            continue
    messagebox.showerror("Chyba", f"Nepodařilo se načíst listy ze souboru:\n{os.path.basename(cesta)}")
    return []


def nacist_sloupce(cesta, sheet_name, header_row):
    df = nacist_excel(cesta, sheet_name=sheet_name, header=header_row, nrows=0)
    return list(df.columns) if df is not None else []


def ulozit_excel_s_formatem(cesta_puvodni, cesta_vystup, df, sheet_name, header_row_0based):
    """Zkopíruje originál Excelu i s barvami/styly a přepíše pouze hodnoty buněk."""
    if os.path.abspath(cesta_puvodni) != os.path.abspath(cesta_vystup):
        shutil.copyfile(cesta_puvodni, cesta_vystup)

    wb = openpyxl.load_workbook(cesta_vystup)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    hdr_row_excel = header_row_0based + 1
    data_start_row = hdr_row_excel + 1

    excel_cols = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=hdr_row_excel, column=col_idx).value
        if val is not None:
            excel_cols[str(val).strip()] = col_idx

    next_free_col = ws.max_column + 1
    for col_name in df.columns:
        col_key = str(col_name).strip()
        if col_key not in excel_cols:
            ws.cell(row=hdr_row_excel, column=next_free_col, value=col_name)
            excel_cols[col_key] = next_free_col
            next_free_col += 1

    for r_idx, row in df.reset_index(drop=True).iterrows():
        excel_r = data_start_row + r_idx
        for col_name in df.columns:
            col_key = str(col_name).strip()
            c_idx = excel_cols[col_key]
            val = row[col_name]
            ws.cell(row=excel_r, column=c_idx, value=None if pd.isna(val) else val)

    max_df_row = data_start_row + len(df) - 1
    if ws.max_row > max_df_row:
        for r in range(max_df_row + 1, ws.max_row + 1):
            for c_idx in excel_cols.values():
                ws.cell(row=r, column=c_idx, value=None)

    wb.save(cesta_vystup)
    wb.close()


# ================= STRÁNKA 1: SPRÁVA SOUBORŮ =================
def pridat_soubor_ui():
    box = ttk.LabelFrame(scroll_frame_soubory, text=f" Soubor {len(seznam_souboru_ui) + 1} ", padding=10)
    box.pack(fill=tk.X, expand=True, pady=5, padx=5)

    f1 = ttk.Frame(box)
    f1.pack(fill=tk.X, pady=(0, 5))

    btn_browse = ttk.Button(f1, text="Procházet...")
    btn_browse.pack(side=tk.LEFT, padx=(0, 10))

    lbl_path = ttk.Label(f1, text="Není vybráno", font=("Segoe UI", 8, "italic"), foreground="gray")
    lbl_path.pack(side=tk.LEFT, fill=tk.X, expand=True)

    polozka = {
        "box": box, "cesta": "", "lbl_path": lbl_path,
        "entry_alias": None, "cb_list": None, "sp_hdr": None
    }

    btn_del = ttk.Button(f1, text="✕", width=3, command=lambda: odebrat_soubor_ui(polozka))
    btn_del.pack(side=tk.RIGHT)

    f2 = ttk.Frame(box)
    f2.pack(fill=tk.X, pady=(5, 0))

    ttk.Label(f2, text="Název/Alias:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
    entry_alias = ttk.Entry(f2, width=22)
    entry_alias.grid(row=0, column=1, padx=(0, 15), sticky=tk.W)

    ttk.Label(f2, text="List:").grid(row=0, column=2, sticky=tk.W, padx=(0, 5))
    cb_list = ttk.Combobox(f2, width=20, state="readonly")
    cb_list.grid(row=0, column=3, padx=(0, 15), sticky=tk.W)

    ttk.Label(f2, text="Řádek záhlaví:").grid(row=0, column=4, sticky=tk.W, padx=(0, 5))
    sp_hdr = ttk.Spinbox(f2, from_=1, to=100, width=5)
    sp_hdr.set(1)
    sp_hdr.grid(row=0, column=5, sticky=tk.W)

    polozka["entry_alias"] = entry_alias
    polozka["cb_list"] = cb_list
    polozka["sp_hdr"] = sp_hdr

    def vybrat():
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls *.xlsm")])
        if path:
            polozka["cesta"] = path
            lbl_path.config(text=os.path.basename(path), foreground="#2e7d32")
            entry_alias.delete(0, tk.END)
            entry_alias.insert(0, os.path.splitext(os.path.basename(path))[0])
            listy = nacist_listy(path)
            cb_list["values"] = listy
            if listy:
                cb_list.set(listy[0])

    btn_browse.config(command=vybrat)
    seznam_souboru_ui.append(polozka)

    canvas_soubory.update_idletasks()
    canvas_soubory.configure(scrollregion=canvas_soubory.bbox("all"))
    aktualizovat_vystupni_sekci()


def odebrat_soubor_ui(polozka):
    polozka["box"].destroy()
    seznam_souboru_ui.remove(polozka)
    for idx, item in enumerate(seznam_souboru_ui):
        item["box"].config(text=f" Soubor {idx + 1} ")
    canvas_soubory.update_idletasks()
    canvas_soubory.configure(scrollregion=canvas_soubory.bbox("all"))
    aktualizovat_vystupni_sekci()


def vybrat_cilovy_soubor():
    global cesta_cil_global
    path = filedialog.asksaveasfilename(filetypes=[("Excel (.xlsx)", "*.xlsx")], defaultextension=".xlsx")
    if path:
        cesta_cil_global = path
        lbl_cil_path.config(text=os.path.basename(path), foreground="#2e7d32")


def aktualizovat_vystupni_sekci():
    if len(seznam_souboru_ui) <= 1:
        box_vystup_multi.pack_forget()
        box_vystup_single.pack(fill=tk.X, pady=(10, 0))
    else:
        box_vystup_single.pack_forget()
        box_vystup_multi.pack(fill=tk.X, pady=(10, 0))


def prejit_na_pracovni_plochu():
    global nactene_soubory
    if not seznam_souboru_ui:
        messagebox.showwarning("Chyba", "Musíš přidat alespoň jeden soubor!")
        return

    nactene_soubory.clear()
    for idx, item in enumerate(seznam_souboru_ui):
        cesta = item["cesta"]
        alias = item["entry_alias"].get().strip() or f"Soubor_{idx + 1}"
        sheet = item["cb_list"].get()
        try:
            hdr = int(item["sp_hdr"].get()) - 1
        except ValueError:
            hdr = 0

        if not cesta:
            messagebox.showwarning("Chyba", f"Soubor {idx + 1} nemá vybranou cestu!")
            return

        sloupce = nacist_sloupce(cesta, sheet, hdr)
        if not sloupce:
            messagebox.showerror("Chyba", f"Nepodařilo se načíst sloupce ze souboru '{alias}'.")
            return

        nactene_soubory[alias] = {"cesta": cesta, "sheet": sheet, "hdr": hdr, "sloupce": sloupce}

    if len(nactene_soubory) > 1 and not cesta_cil_global:
        messagebox.showwarning("Chyba", "Při práci s více soubory musíš určit výstupní soubor!")
        return

    frame_strana1.pack_forget()
    frame_strana2.pack(fill=tk.BOTH, expand=True)

    if not seznam_operaci:
        pridat_radek_operace()


# ================= POP-UP OKNA PRO JEDNOTLIVÉ KROKY =================

# 1. KOPÍROVAT SLOUPCE (1 KÓD)
def popup_kopirovat_1klic(data_radku):
    popup = tk.Toplevel(root)
    popup.title("Kopírovat sloupce (1 kód)")
    popup.geometry("750x640")
    popup.grab_set()

    f_help = ttk.Frame(popup, padding=(12, 12, 12, 0))
    f_help.pack(fill=tk.X)
    pridat_napovedu(f_help, "Tato funkce najde v Cílovém souboru řádky, které mají stejný kód (P/N) jako ve Zdrojovém souboru, a překopíruje z nich hodnoty ze zadaných sloupců.")

    f_top = ttk.Frame(popup, padding=12)
    f_top.pack(fill=tk.X)
    seznam_aliasu = list(nactene_soubory.keys())

    ttk.Label(f_top, text="Zdrojový soubor:", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
    cb_src_file = ttk.Combobox(f_top, values=seznam_aliasu, width=28, state="readonly")
    cb_src_file.set(data_radku.get("src_file", seznam_aliasu[0]))
    cb_src_file.grid(row=0, column=1, padx=5, pady=2)

    ttk.Label(f_top, text="Cílový soubor:", font=("Segoe UI", 9, "bold")).grid(row=0, column=2, sticky=tk.W, padx=(20, 5), pady=2)
    cb_dst_file = ttk.Combobox(f_top, values=seznam_aliasu, width=28, state="readonly")
    cb_dst_file.set(data_radku.get("dst_file", seznam_aliasu[-1] if len(seznam_aliasu) > 1 else seznam_aliasu[0]))
    cb_dst_file.grid(row=0, column=3, padx=5, pady=2)

    ttk.Label(f_top, text="Kód ve Zdrojovém s.:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
    cb_klic_src = ttk.Combobox(f_top, width=28, state="readonly")
    cb_klic_src.grid(row=1, column=1, padx=5, pady=2)

    ttk.Label(f_top, text="Kód v Cílovém s.:").grid(row=1, column=2, sticky=tk.W, padx=(20, 5), pady=2)
    cb_klic_dst = ttk.Combobox(f_top, width=28, state="readonly")
    cb_klic_dst.grid(row=1, column=3, padx=5, pady=2)

    ttk.Label(f_top, text="Konec dat podle (Zdroj):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=(6, 2))
    cb_konec_col = ttk.Combobox(f_top, width=28, state="readonly")
    cb_konec_col.grid(row=2, column=1, padx=5, pady=(6, 2))

    def refresh_cols(e=None):
        sc = nactene_soubory.get(cb_src_file.get(), {}).get("sloupce", [])
        dc = nactene_soubory.get(cb_dst_file.get(), {}).get("sloupce", [])
        cb_klic_src["values"] = sc
        if sc:
            cb_klic_src.set(data_radku.get("klic_src", sc[0]))
        cb_klic_dst["values"] = dc
        if dc:
            cb_klic_dst.set(data_radku.get("klic_dst", dc[0]))
        cb_konec_col["values"] = sc
        if sc:
            cb_konec_col.set(data_radku.get("konec_col", sc[0]))

    cb_src_file.bind("<<ComboboxSelected>>", refresh_cols)
    cb_dst_file.bind("<<ComboboxSelected>>", refresh_cols)
    refresh_cols()

    canvas = tk.Canvas(popup, borderwidth=0, highlightthickness=0)
    scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
    scroll_frame = ttk.Frame(canvas)
    scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    pop_win_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(pop_win_id, width=e.width))
    canvas.configure(yscrollcommand=scrollbar.set)
    nastavit_scrolovani(canvas)
    canvas.pack(side="top", fill="both", expand=True, padx=10, pady=5)
    scrollbar.pack(side="right", fill="y")

    ui_prvky = []

    def pridat_dvojici(src_val=None, dst_val=None):
        f = ttk.LabelFrame(scroll_frame, text=f" Pravidlo {len(ui_prvky) + 1} ", padding=8)
        f.pack(fill=tk.X, expand=True, pady=4, padx=5)
        sc = nactene_soubory.get(cb_src_file.get(), {}).get("sloupce", [])
        dc = nactene_soubory.get(cb_dst_file.get(), {}).get("sloupce", [])
        ttk.Label(f, text="Zdrojový sloupec:").grid(row=0, column=0, sticky=tk.W, padx=5)
        cb_s = ttk.Combobox(f, values=sc, width=32, state="readonly")
        cb_s.set(src_val or (sc[0] if sc else ""))
        cb_s.grid(row=0, column=1, padx=5, pady=2)
        ttk.Label(f, text="Cílový sloupec:").grid(row=1, column=0, sticky=tk.W, padx=5)
        cb_d = ttk.Combobox(f, values=dc, width=32, state="readonly")
        cb_d.set(dst_val or (dc[0] if dc else ""))
        cb_d.grid(row=1, column=1, padx=5, pady=2)
        p = {"frame": f, "src": cb_s, "dst": cb_d}
        ttk.Button(f, text="✕", width=3, command=lambda: (f.destroy(), ui_prvky.remove(p))).grid(row=0, column=2, rowspan=2, padx=10)
        ui_prvky.append(p)
        canvas.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

    if data_radku.get("mapovani_rules"):
        for r in data_radku["mapovani_rules"]:
            pridat_dvojici(r["src"], r["dst"])
    else:
        pridat_dvojici()

    bot_frame = ttk.Frame(popup, padding=10)
    bot_frame.pack(fill=tk.X, side=tk.BOTTOM)
    ttk.Button(bot_frame, text="+ Přidat další pravidlo", command=pridat_dvojici).pack(anchor=tk.W, pady=(0, 10))

    def ulozit():
        data_radku.update({
            "src_file": cb_src_file.get(), "dst_file": cb_dst_file.get(),
            "klic_src": cb_klic_src.get(), "klic_dst": cb_klic_dst.get(), "konec_col": cb_konec_col.get(),
            "mapovani_rules": [{"src": x["src"].get(), "dst": x["dst"].get()} for x in ui_prvky]
        })
        data_radku["btn_upravit"].config(text=f"⚙ Mapování [{cb_src_file.get()}] ➔ [{cb_dst_file.get()}] ({len(ui_prvky)})")
        popup.destroy()

    ttk.Button(bot_frame, text="Uložit mapování", command=ulozit).pack(fill=tk.X, ipady=4)


# 2. KOPÍROVAT SLOUPCE (2 KÓDY)
def popup_kopirovat_2klice(data_radku):
    popup = tk.Toplevel(root)
    popup.title("Kopírovat sloupce (2 kódy)")
    popup.geometry("750x680")
    popup.grab_set()

    f_help = ttk.Frame(popup, padding=(12, 12, 12, 0))
    f_help.pack(fill=tk.X)
    pridat_napovedu(f_help, "Páruje řádky podle kombinace DVOU klíčů současně (např. P/N kód + Typ). Data překopíruje pouze do řádků, kde se shodují oba klíče najednou.")

    f_top = ttk.Frame(popup, padding=12)
    f_top.pack(fill=tk.X)
    seznam_aliasu = list(nactene_soubory.keys())

    ttk.Label(f_top, text="Zdrojový soubor:", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
    cb_src_file = ttk.Combobox(f_top, values=seznam_aliasu, width=28, state="readonly")
    cb_src_file.set(data_radku.get("src_file", seznam_aliasu[0]))
    cb_src_file.grid(row=0, column=1, padx=5, pady=2)

    ttk.Label(f_top, text="Cílový soubor:", font=("Segoe UI", 9, "bold")).grid(row=0, column=2, sticky=tk.W, padx=(20, 5), pady=2)
    cb_dst_file = ttk.Combobox(f_top, values=seznam_aliasu, width=28, state="readonly")
    cb_dst_file.set(data_radku.get("dst_file", seznam_aliasu[-1] if len(seznam_aliasu) > 1 else seznam_aliasu[0]))
    cb_dst_file.grid(row=0, column=3, padx=5, pady=2)

    ttk.Label(f_top, text="Klíč 1 (Zdroj):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
    cb_k1_src = ttk.Combobox(f_top, width=28, state="readonly")
    cb_k1_src.grid(row=1, column=1, padx=5, pady=2)
    ttk.Label(f_top, text="Klíč 1 (Cíl):").grid(row=1, column=2, sticky=tk.W, padx=(20, 5), pady=2)
    cb_k1_dst = ttk.Combobox(f_top, width=28, state="readonly")
    cb_k1_dst.grid(row=1, column=3, padx=5, pady=2)

    ttk.Label(f_top, text="Klíč 2 (Zdroj):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
    cb_k2_src = ttk.Combobox(f_top, width=28, state="readonly")
    cb_k2_src.grid(row=2, column=1, padx=5, pady=2)
    ttk.Label(f_top, text="Klíč 2 (Cíl):").grid(row=2, column=2, sticky=tk.W, padx=(20, 5), pady=2)
    cb_k2_dst = ttk.Combobox(f_top, width=28, state="readonly")
    cb_k2_dst.grid(row=2, column=3, padx=5, pady=2)

    ttk.Label(f_top, text="Konec dat podle (Zdroj):").grid(row=3, column=0, sticky=tk.W, padx=5, pady=(6, 2))
    cb_konec_col = ttk.Combobox(f_top, width=28, state="readonly")
    cb_konec_col.grid(row=3, column=1, padx=5, pady=(6, 2))

    def refresh_cols(e=None):
        sc = nactene_soubory.get(cb_src_file.get(), {}).get("sloupce", [])
        dc = nactene_soubory.get(cb_dst_file.get(), {}).get("sloupce", [])
        cb_k1_src["values"] = sc
        if sc:
            cb_k1_src.set(data_radku.get("k1_src", sc[0]))
        cb_k1_dst["values"] = dc
        if dc:
            cb_k1_dst.set(data_radku.get("k1_dst", dc[0]))
        cb_k2_src["values"] = sc
        if sc:
            cb_k2_src.set(data_radku.get("k2_src", sc[0]))
        cb_k2_dst["values"] = dc
        if dc:
            cb_k2_dst.set(data_radku.get("k2_dst", dc[0]))
        cb_konec_col["values"] = sc
        if sc:
            cb_konec_col.set(data_radku.get("konec_col", sc[0]))

    cb_src_file.bind("<<ComboboxSelected>>", refresh_cols)
    cb_dst_file.bind("<<ComboboxSelected>>", refresh_cols)
    refresh_cols()

    canvas = tk.Canvas(popup, borderwidth=0, highlightthickness=0)
    scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
    scroll_frame = ttk.Frame(canvas)
    scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    pop_win_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(pop_win_id, width=e.width))
    canvas.configure(yscrollcommand=scrollbar.set)
    nastavit_scrolovani(canvas)
    canvas.pack(side="top", fill="both", expand=True, padx=10, pady=5)
    scrollbar.pack(side="right", fill="y")

    ui_prvky = []

    def pridat_dvojici(src_val=None, dst_val=None):
        f = ttk.LabelFrame(scroll_frame, text=f" Pravidlo {len(ui_prvky) + 1} ", padding=8)
        f.pack(fill=tk.X, expand=True, pady=4, padx=5)
        sc = nactene_soubory.get(cb_src_file.get(), {}).get("sloupce", [])
        dc = nactene_soubory.get(cb_dst_file.get(), {}).get("sloupce", [])
        ttk.Label(f, text="Zdrojový sloupec:").grid(row=0, column=0, sticky=tk.W, padx=5)
        cb_s = ttk.Combobox(f, values=sc, width=32, state="readonly")
        cb_s.set(src_val or (sc[0] if sc else ""))
        cb_s.grid(row=0, column=1, padx=5, pady=2)
        ttk.Label(f, text="Cílový sloupec:").grid(row=1, column=0, sticky=tk.W, padx=5)
        cb_d = ttk.Combobox(f, values=dc, width=32, state="readonly")
        cb_d.set(dst_val or (dc[0] if dc else ""))
        cb_d.grid(row=1, column=1, padx=5, pady=2)
        p = {"frame": f, "src": cb_s, "dst": cb_d}
        ttk.Button(f, text="✕", width=3, command=lambda: (f.destroy(), ui_prvky.remove(p))).grid(row=0, column=2, rowspan=2, padx=10)
        ui_prvky.append(p)
        canvas.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

    if data_radku.get("mapovani_rules"):
        for r in data_radku["mapovani_rules"]:
            pridat_dvojici(r["src"], r["dst"])
    else:
        pridat_dvojici()

    bot_frame = ttk.Frame(popup, padding=10)
    bot_frame.pack(fill=tk.X, side=tk.BOTTOM)
    ttk.Button(bot_frame, text="+ Přidat další pravidlo", command=pridat_dvojici).pack(anchor=tk.W, pady=(0, 10))

    def ulozit():
        data_radku.update({
            "src_file": cb_src_file.get(), "dst_file": cb_dst_file.get(),
            "k1_src": cb_k1_src.get(), "k1_dst": cb_k1_dst.get(),
            "k2_src": cb_k2_src.get(), "k2_dst": cb_k2_dst.get(), "konec_col": cb_konec_col.get(),
            "mapovani_rules": [{"src": x["src"].get(), "dst": x["dst"].get()} for x in ui_prvky]
        })
        data_radku["btn_upravit"].config(text=f"⚙ [2 Klíče] [{cb_src_file.get()}] ➔ [{cb_dst_file.get()}]")
        popup.destroy()

    ttk.Button(bot_frame, text="Uložit mapování", command=ulozit).pack(fill=tk.X, ipady=4)


# 3. ROZDĚLIT DATA NA IAM / OE / OES
def popup_rozdeleni_iam_oe(data_radku):
    popup = tk.Toplevel(root)
    popup.title("Rozdělit data na IAM / OE / OES")
    popup.geometry("680x680")
    popup.grab_set()

    f = ttk.Frame(popup, padding=15)
    f.pack(fill=tk.BOTH, expand=True)
    pridat_napovedu(f, "Sečte zadané sloupce (např. pololetí). Pokud příjemce obsahuje přesnou zkratku (např. '/MA'), vloží 100 % do IAM. Jinak rozpočítá zadaná procenta mezi OE a OES.")

    seznam_aliasu = list(nactene_soubory.keys())

    box1 = ttk.LabelFrame(f, text=" 1. Soubory a Párovací P/N kód ", padding=10)
    box1.pack(fill=tk.X, pady=(0, 10))
    ttk.Label(box1, text="Zdrojový s.:").grid(row=0, column=0, sticky=tk.W)
    cb_src = ttk.Combobox(box1, values=seznam_aliasu, width=20, state="readonly")
    cb_src.set(data_radku.get("src_file", seznam_aliasu[0]))
    cb_src.grid(row=0, column=1, padx=5, pady=2)
    ttk.Label(box1, text="Cílový s.:").grid(row=0, column=2, sticky=tk.W, padx=(15, 0))
    cb_dst = ttk.Combobox(box1, values=seznam_aliasu, width=20, state="readonly")
    cb_dst.set(data_radku.get("dst_file", seznam_aliasu[-1] if len(seznam_aliasu) > 1 else seznam_aliasu[0]))
    cb_dst.grid(row=0, column=3, padx=5, pady=2)
    ttk.Label(box1, text="P/N Zdroj:").grid(row=1, column=0, sticky=tk.W)
    cb_pn_src = ttk.Combobox(box1, width=20, state="readonly")
    cb_pn_src.grid(row=1, column=1, padx=5, pady=2)
    ttk.Label(box1, text="P/N Cíl:").grid(row=1, column=2, sticky=tk.W, padx=(15, 0))
    cb_pn_dst = ttk.Combobox(box1, width=20, state="readonly")
    cb_pn_dst.grid(row=1, column=3, padx=5, pady=2)

    box2 = ttk.LabelFrame(f, text=" 2. Vstupní data u Zákazníka ", padding=10)
    box2.pack(fill=tk.X, pady=(0, 10))
    ttk.Label(box2, text="Sloupec s příjemcem:").grid(row=0, column=0, sticky=tk.W)
    cb_rec_col = ttk.Combobox(box2, width=20, state="readonly")
    cb_rec_col.grid(row=0, column=1, padx=5, pady=2)
    ttk.Label(box2, text="Hledaný text (interní):").grid(row=0, column=2, sticky=tk.W, padx=(15, 0))
    txt_filter = ttk.Entry(box2, width=15)
    txt_filter.insert(0, data_radku.get("filter_text", "/MA"))
    txt_filter.grid(row=0, column=3, padx=5, pady=2)
    ttk.Label(box2, text="Sčítané sloupce u Zákazníka (např. pololetí):").grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(5, 2))
    cb_sum1 = ttk.Combobox(box2, width=20, state="readonly")
    cb_sum1.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=5)
    cb_sum2 = ttk.Combobox(box2, width=20, state="readonly")
    cb_sum2.grid(row=2, column=2, columnspan=2, sticky=tk.W, padx=5)

    box3 = ttk.LabelFrame(f, text=" 3. Cílové umístění v Naší tabulce ", padding=10)
    box3.pack(fill=tk.X, pady=(0, 10))
    ttk.Label(box3, text="Sloupec s kategorií (např. Sloupec S):").grid(row=0, column=0, sticky=tk.W)
    cb_cat_col = ttk.Combobox(box3, width=22, state="readonly")
    cb_cat_col.grid(row=0, column=1, padx=5, pady=2)
    ttk.Label(box3, text="Hodnotový sloupec (např. Rok 2026):").grid(row=1, column=0, sticky=tk.W)
    cb_val_col = ttk.Combobox(box3, width=22, state="readonly")
    cb_val_col.grid(row=1, column=1, padx=5, pady=2)

    box4 = ttk.LabelFrame(f, text=" 4. Pravidla procentuálního rozdělení ", padding=10)
    box4.pack(fill=tk.X, pady=(0, 10))
    ttk.Label(box4, text="Když obsahuje hledaný text ➔ do kategorie:", font=("Segoe UI", 8, "bold")).grid(row=0, column=0, columnspan=2, sticky=tk.W)
    txt_m_cat = ttk.Entry(box4, width=10)
    txt_m_cat.insert(0, data_radku.get("m_cat", "IAM"))
    txt_m_cat.grid(row=0, column=2, padx=2)
    ttk.Label(box4, text="Procento:").grid(row=0, column=3, sticky=tk.W)
    txt_m_pct = ttk.Entry(box4, width=6)
    txt_m_pct.insert(0, data_radku.get("m_pct", "100"))
    txt_m_pct.grid(row=0, column=4, padx=2)

    ttk.Label(box4, text="Jinak (Ostatní) ➔ Rozdělit do kategorií:", font=("Segoe UI", 8, "bold")).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))
    txt_e1_cat = ttk.Entry(box4, width=10)
    txt_e1_cat.insert(0, data_radku.get("e1_cat", "OE"))
    txt_e1_cat.grid(row=1, column=2, padx=2)
    txt_e1_pct = ttk.Entry(box4, width=6)
    txt_e1_pct.insert(0, data_radku.get("e1_pct", "30"))
    txt_e1_pct.grid(row=1, column=3, padx=2)
    txt_e2_cat = ttk.Entry(box4, width=10)
    txt_e2_cat.insert(0, data_radku.get("e2_cat", "OES"))
    txt_e2_cat.grid(row=1, column=4, padx=2)
    txt_e2_pct = ttk.Entry(box4, width=6)
    txt_e2_pct.insert(0, data_radku.get("e2_pct", "70"))
    txt_e2_pct.grid(row=1, column=5, padx=2)

    def refresh_cols(e=None):
        sc = nactene_soubory.get(cb_src.get(), {}).get("sloupce", [])
        dc = nactene_soubory.get(cb_dst.get(), {}).get("sloupce", [])
        cb_pn_src["values"] = sc
        if sc:
            cb_pn_src.set(data_radku.get("pn_src", sc[0]))
        cb_pn_dst["values"] = dc
        if dc:
            cb_pn_dst.set(data_radku.get("pn_dst", dc[0]))
        cb_rec_col["values"] = sc
        if sc:
            cb_rec_col.set(data_radku.get("rec_col", sc[0]))
        cb_sum1["values"] = sc
        if sc:
            cb_sum1.set(data_radku.get("sum1", sc[0]))
        cb_sum2["values"] = ["-- Žádný --"] + sc
        if sc:
            cb_sum2.set(data_radku.get("sum2", sc[1] if len(sc) > 1 else "-- Žádný --"))
        cb_cat_col["values"] = dc
        if dc:
            cb_cat_col.set(data_radku.get("cat_col", dc[0]))
        cb_val_col["values"] = dc
        if dc:
            cb_val_col.set(data_radku.get("val_col", dc[0]))

    cb_src.bind("<<ComboboxSelected>>", refresh_cols)
    cb_dst.bind("<<ComboboxSelected>>", refresh_cols)
    refresh_cols()

    def ulozit():
        data_radku.update({
            "src_file": cb_src.get(), "dst_file": cb_dst.get(),
            "pn_src": cb_pn_src.get(), "pn_dst": cb_pn_dst.get(),
            "rec_col": cb_rec_col.get(), "filter_text": txt_filter.get(),
            "sum1": cb_sum1.get(), "sum2": cb_sum2.get(),
            "cat_col": cb_cat_col.get(), "val_col": cb_val_col.get(),
            "m_cat": txt_m_cat.get(), "m_pct": txt_m_pct.get(),
            "e1_cat": txt_e1_cat.get(), "e1_pct": txt_e1_pct.get(),
            "e2_cat": txt_e2_cat.get(), "e2_pct": txt_e2_pct.get()
        })
        data_radku["btn_upravit"].config(text=f"⚙ [Rozdělit %] {cb_src.get()} ➔ {cb_dst.get()} [{cb_val_col.get()}]")
        popup.destroy()

    ttk.Button(f, text="Uložit nastavení", command=ulozit).pack(side=tk.BOTTOM, fill=tk.X, ipady=4)


# 4. SEČÍST SLOUPCE DO NOVÉHO
def popup_secist_sloupce(data_radku):
    popup = tk.Toplevel(root)
    popup.title("Sečíst sloupce do nového")
    popup.geometry("520x500")
    popup.grab_set()

    f = ttk.Frame(popup, padding=15)
    f.pack(fill=tk.BOTH, expand=True)
    pridat_napovedu(f, "Sečte hodnoty ze zaškrtnutých sloupců řádek po řádku a vytvoří nový sloupec s celkovým součtem.")

    seznam_aliasu = list(nactene_soubory.keys())

    ttk.Label(f, text="1. Vyber soubor:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 2))
    cb_target_file = ttk.Combobox(f, values=seznam_aliasu, width=42, state="readonly")
    cb_target_file.set(data_radku.get("target_file", seznam_aliasu[0]))
    cb_target_file.pack(anchor=tk.W, pady=(0, 10))

    ttk.Label(f, text="2. Název nového sloupce pro součet:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 2))
    txt_new_col = ttk.Entry(f, width=45)
    txt_new_col.insert(0, data_radku.get("new_col_name", "Součet_pololetí"))
    txt_new_col.pack(anchor=tk.W, pady=(0, 10))

    ttk.Label(f, text="3. Vyber sloupce k sečtení:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 5))
    frame_cols = ttk.Frame(f)
    frame_cols.pack(fill=tk.BOTH, expand=True, pady=5)
    check_vars = {}

    def refresh_cols(e=None):
        for child in frame_cols.winfo_children():
            child.destroy()
        check_vars.clear()
        cols = nactene_soubory.get(cb_target_file.get(), {}).get("sloupce", [])
        saved_sum = data_radku.get("sum_cols", [])
        for c in cols:
            var = tk.BooleanVar(value=(c in saved_sum))
            ttk.Checkbutton(frame_cols, text=c, variable=var).pack(anchor=tk.W, pady=1)
            check_vars[c] = var

    cb_target_file.bind("<<ComboboxSelected>>", refresh_cols)
    refresh_cols()

    def ulozit():
        selected = [col for col, var in check_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("Chyba", "Musíš vybrat alespoň jeden sloupec!")
            return
        data_radku.update({"target_file": cb_target_file.get(), "new_col_name": txt_new_col.get().strip(), "sum_cols": selected})
        data_radku["btn_upravit"].config(text=f"⚙ [{cb_target_file.get()}] Sečíst ({len(selected)} sloupců) ➔ [{txt_new_col.get()}]")
        popup.destroy()

    ttk.Button(f, text="Uložit", command=ulozit).pack(side=tk.BOTTOM, fill=tk.X, ipady=4)


# 5. SPOJIT TEXT ZE 2 SLOUPCŮ
def popup_spojit_text(data_radku):
    popup = tk.Toplevel(root)
    popup.title("Spojit text ze 2 sloupců")
    popup.geometry("520x460")
    popup.grab_set()

    f = ttk.Frame(popup, padding=15)
    f.pack(fill=tk.BOTH, expand=True)
    pridat_napovedu(f, "Připojí text ze Zdrojového sloupce k Cílovému sloupci. Mezi ně vloží zadaný text a dvojtečku ('TextA' + ' [Vložený text]: ' + 'TextB').")

    seznam_aliasu = list(nactene_soubory.keys())

    ttk.Label(f, text="1. Vyber upravovaný soubor:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 2))
    cb_target_file = ttk.Combobox(f, values=seznam_aliasu, width=45, state="readonly")
    cb_target_file.set(data_radku.get("target_file", seznam_aliasu[0]))
    cb_target_file.pack(anchor=tk.W, pady=(0, 10))

    ttk.Label(f, text="2. Cílový sloupec (kam se uloží výsledek):", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
    cb_dst_col = ttk.Combobox(f, width=45, state="readonly")
    cb_dst_col.pack(anchor=tk.W, pady=(0, 8))

    ttk.Label(f, text="3. Vložený text / Předpona (mezi texty):", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
    txt_prefix = ttk.Entry(f, width=48)
    txt_prefix.insert(0, data_radku.get("vlozeny_text", "kategorie"))
    txt_prefix.pack(anchor=tk.W, pady=(0, 8))

    ttk.Label(f, text="4. Zdrojový sloupec (text na konec):", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
    cb_src_col = ttk.Combobox(f, width=45, state="readonly")
    cb_src_col.pack(anchor=tk.W, pady=(0, 8))

    lbl_preview = ttk.Label(f, text="", font=("Segoe UI", 9, "italic"), foreground="#2e7d32", wraplength=460, justify="left")
    lbl_preview.pack(anchor=tk.W, pady=(5, 10))

    def update_preview(e=None):
        lbl_preview.config(text=f"💡 Náhled výsledku:\n\"Tohle je text {txt_prefix.get()}: tohle je taky text\"")

    def refresh_cols(e=None):
        cols = nactene_soubory.get(cb_target_file.get(), {}).get("sloupce", [])
        cb_dst_col["values"] = cols
        cb_src_col["values"] = cols
        if cols:
            cb_dst_col.set(data_radku.get("klic1", cols[0]))
            cb_src_col.set(data_radku.get("klic2", cols[1] if len(cols) > 1 else cols[0]))
        update_preview()

    cb_target_file.bind("<<ComboboxSelected>>", refresh_cols)
    cb_dst_col.bind("<<ComboboxSelected>>", update_preview)
    cb_src_col.bind("<<ComboboxSelected>>", update_preview)
    txt_prefix.bind("<KeyRelease>", update_preview)
    refresh_cols()

    def ulozit():
        data_radku.update({"target_file": cb_target_file.get(), "klic1": cb_dst_col.get(), "klic2": cb_src_col.get(), "vlozeny_text": txt_prefix.get()})
        data_radku["btn_upravit"].config(text=f"⚙ [{cb_target_file.get()}] Spojit text [{cb_dst_col.get()}] + [{cb_src_col.get()}]")
        popup.destroy()

    ttk.Button(f, text="Uložit", command=ulozit).pack(side=tk.BOTTOM, fill=tk.X, ipady=4)


# 6. OSTATNÍ PROSTÉ OPERACE (SLOUČIT / PŘIČÍST / SMAZAT / VLOŽIT)
def otevrit_nastaveni_kroku(data_radku):
    op = data_radku["operace"].get()

    if op == "Rozdělit data na IAM / OE / OES":
        popup_rozdeleni_iam_oe(data_radku)
        return
    elif op == "Sečíst sloupce do nového":
        popup_secist_sloupce(data_radku)
        return
    elif op == "Kopírovat sloupce (2 kódy)":
        popup_kopirovat_2klice(data_radku)
        return
    elif op == "Spojit text ze 2 sloupců":
        popup_spojit_text(data_radku)
        return

    popup = tk.Toplevel(root)
    popup.title(f"Nastavení: {op}")
    popup.geometry("500x440")
    popup.grab_set()

    main = ttk.Frame(popup, padding=15)
    main.pack(fill=tk.BOTH, expand=True)

    if op == "Sloučit duplicitní řádky":
        pridat_napovedu(main, "Sloučí duplicitní řádky podle zvoleného sloupce. Čísla se sečtou, texty zůstanou první.")
    elif op == "Přičíst sloupec k jinému":
        pridat_napovedu(main, "Přičte číselné hodnoty ze zdrojového sloupce k cílovému sloupci.")
    elif op == "Smazat obsah sloupce":
        pridat_napovedu(main, "Vyprázdní veškeré hodnoty v celém vybraném sloupci.")
    elif op == "Vložit stejný text do sloupce":
        pridat_napovedu(main, "Vloží zadaný text do vybraného sloupce (lze omezit delkou jiného sloupce).")

    seznam_aliasu = list(nactene_soubory.keys())
    ttk.Label(main, text="1. Vyber upravovaný soubor:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 2))
    cb_target_file = ttk.Combobox(main, values=seznam_aliasu, width=45, state="readonly")
    cb_target_file.set(data_radku.get("target_file", seznam_aliasu[0]))
    cb_target_file.pack(anchor=tk.W, pady=(0, 12))

    dyn = ttk.Frame(main)
    dyn.pack(fill=tk.BOTH, expand=True)

    def nacist_cols():
        return nactene_soubory.get(cb_target_file.get(), {}).get("sloupce", [])

    if op == "Sloučit duplicitní řádky":
        ttk.Label(dyn, text="Sloučit duplicitní řádky podle sloupce:", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
        cb_col = ttk.Combobox(dyn, width=45, state="readonly")
        cb_col.pack(anchor=tk.W, pady=(0, 10))

        def ref(e=None):
            cols = nacist_cols()
            cb_col["values"] = cols
            if cols:
                cb_col.set(data_radku.get("klic1", cols[0]))

        cb_target_file.bind("<<ComboboxSelected>>", ref)
        ref()

        def ulozit():
            data_radku.update({"target_file": cb_target_file.get(), "klic1": cb_col.get()})
            data_radku["btn_upravit"].config(text=f"⚙ [{cb_target_file.get()}] Sloučit podle [{cb_col.get()}]")
            popup.destroy()

    elif op == "Přičíst sloupec k jinému":
        ttk.Label(dyn, text="Zdrojový sloupec:", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
        cb_s = ttk.Combobox(dyn, width=45, state="readonly")
        cb_s.pack(anchor=tk.W, pady=(0, 8))
        ttk.Label(dyn, text="Cílový sloupec (kam přičíst):", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
        cb_r = ttk.Combobox(dyn, width=45, state="readonly")
        cb_r.pack(anchor=tk.W, pady=(0, 8))
        lbl_note = ttk.Label(dyn, text="", font=("Segoe UI", 9, "italic"), foreground="#2e7d32")
        lbl_note.pack(anchor=tk.W, pady=(5, 10))

        def ref(e=None):
            cols = nacist_cols()
            cb_s["values"] = cols
            cb_r["values"] = cols
            if cols:
                cb_s.set(data_radku.get("klic1", cols[0]))
                cb_r.set(data_radku.get("klic2", cols[0]))
            lbl_note.config(text=f"💡 Výsledný součet bude uložen ve sloupci: '{cb_r.get()}'")

        cb_s.bind("<<ComboboxSelected>>", lambda e: lbl_note.config(text=f"💡 Výsledný součet bude uložen ve sloupci: '{cb_r.get()}'"))
        cb_r.bind("<<ComboboxSelected>>", lambda e: lbl_note.config(text=f"💡 Výsledný součet bude uložen ve sloupci: '{cb_r.get()}'"))
        cb_target_file.bind("<<ComboboxSelected>>", ref)
        ref()

        def ulozit():
            data_radku.update({"target_file": cb_target_file.get(), "klic1": cb_s.get(), "klic2": cb_r.get()})
            data_radku["btn_upravit"].config(text=f"⚙ [{cb_target_file.get()}] Přičíst [{cb_s.get()}] ➔ [{cb_r.get()}]")
            popup.destroy()

    elif op in ["Smazat obsah sloupce", "Vložit stejný text do sloupce"]:
        ttk.Label(dyn, text="Vyber sloupec:", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
        cb_col = ttk.Combobox(dyn, width=45, state="readonly")
        cb_col.pack(anchor=tk.W, pady=(0, 10))
        txt_val, cb_konec = None, None

        if op == "Vložit stejný text do sloupce":
            ttk.Label(dyn, text="Zadej text k vložení:", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
            txt_val = ttk.Entry(dyn, width=48)
            txt_val.insert(0, data_radku.get("hodnota_naplneni", ""))
            txt_val.pack(anchor=tk.W, pady=(0, 10))
            ttk.Label(dyn, text="Konec dat podle sloupce (volitelné):", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 2))
            cb_konec = ttk.Combobox(dyn, width=45, state="readonly")
            cb_konec.pack(anchor=tk.W, pady=(0, 10))

        def ref(e=None):
            cols = nacist_cols()
            cb_col["values"] = cols
            if cols:
                cb_col.set(data_radku.get("vybrany_sloupec", cols[0]))
            if cb_konec:
                cw = ["-- Celý sloupec --"] + cols
                cb_konec["values"] = cw
                cb_konec.set(data_radku.get("konec_col", "-- Celý sloupec --") if data_radku.get("konec_col") in cw else "-- Celý sloupec --")

        cb_target_file.bind("<<ComboboxSelected>>", ref)
        ref()

        def ulozit():
            data_radku.update({"target_file": cb_target_file.get(), "vybrany_sloupec": cb_col.get()})
            if op == "Vložit stejný text do sloupce":
                if txt_val:
                    data_radku["hodnota_naplneni"] = txt_val.get()
                if cb_konec:
                    data_radku["konec_col"] = cb_konec.get()
                data_radku["btn_upravit"].config(text=f"⚙ [{cb_target_file.get()}] Vložit [{cb_col.get()}] = '{txt_val.get()}'")
            else:
                data_radku["btn_upravit"].config(text=f"⚙ [{cb_target_file.get()}] Smazat [{cb_col.get()}]")
            popup.destroy()

    ttk.Button(main, text="Uložit", command=ulozit).pack(side=tk.BOTTOM, fill=tk.X, ipady=4)


# ================= SPRÁVA PROFILŮ (ULOŽIT / NAČÍST) =================
def popup_ulozit_profil():
    if not seznam_operaci:
        messagebox.showwarning("Prázdné operace", "Nemáš vytvořené žádné kroky k uložení!")
        return

    popup = tk.Toplevel(root)
    popup.title("Uložit profil")
    popup.geometry("450x250")
    popup.grab_set()
    f = ttk.Frame(popup, padding=15)
    f.pack(fill=tk.BOTH, expand=True)

    ttk.Label(f, text="Název profilu:", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 2))
    txt_nazev = ttk.Entry(f, width=50)
    txt_nazev.pack(anchor=tk.W, pady=(0, 10))

    ttk.Label(f, text="Popis profilu (volitelné):", font=("Segoe UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 2))
    txt_popis = ttk.Entry(f, width=50)
    txt_popis.pack(anchor=tk.W, pady=(0, 15))

    def potvrdit():
        nazev = txt_nazev.get().strip()
        if not nazev:
            messagebox.showwarning("Chyba", "Musíš zadat název profilu!")
            return

        kroky = []
        for radek in seznam_operaci:
            kroky.append({
                "operace": radek["operace"].get(),
                "src_file": radek.get("src_file", ""), "dst_file": radek.get("dst_file", ""),
                "target_file": radek.get("target_file", ""),
                "klic_src": radek.get("klic_src", ""), "klic_dst": radek.get("klic_dst", ""),
                "k1_src": radek.get("k1_src", ""), "k1_dst": radek.get("k1_dst", ""),
                "k2_src": radek.get("k2_src", ""), "k2_dst": radek.get("k2_dst", ""),
                "konec_col": radek.get("konec_col", ""), "mapovani_rules": radek.get("mapovani_rules", []),
                "klic1": radek.get("klic1", ""), "klic2": radek.get("klic2", ""),
                "vybrany_sloupec": radek.get("vybrany_sloupec", ""), "hodnota_naplneni": radek.get("hodnota_naplneni", ""),
                "vlozeny_text": radek.get("vlozeny_text", ""), "pn_src": radek.get("pn_src", ""),
                "pn_dst": radek.get("pn_dst", ""), "rec_col": radek.get("rec_col", ""),
                "filter_text": radek.get("filter_text", ""), "sum1": radek.get("sum1", ""), "sum2": radek.get("sum2", ""),
                "cat_col": radek.get("cat_col", ""), "val_col": radek.get("val_col", ""),
                "m_cat": radek.get("m_cat", ""), "m_pct": radek.get("m_pct", ""),
                "e1_cat": radek.get("e1_cat", ""), "e1_pct": radek.get("e1_pct", ""),
                "e2_cat": radek.get("e2_cat", ""), "e2_pct": radek.get("e2_pct", ""),
                "new_col_name": radek.get("new_col_name", ""), "sum_cols": radek.get("sum_cols", [])
            })

        profily = nacist_profily()
        profily[nazev] = {"popis": txt_popis.get().strip(), "pouzite_soubory": list(nactene_soubory.keys()), "kroky": kroky}
        ulozit_profily(profily)
        messagebox.showinfo("Hotovo", f"Profil '{nazev}' byl úspěšně uložen!")
        popup.destroy()

    ttk.Button(f, text="Uložit profil", command=potvrdit).pack(side=tk.BOTTOM, fill=tk.X, ipady=4)


def popup_nacist_profil():
    profily = nacist_profily()
    if not profily:
        messagebox.showinfo("Žádné profily", "Zatím nemáš uložené žádné konfigurace profilů.")
        return

    popup = tk.Toplevel(root)
    popup.title("Správa profilů")
    popup.geometry("560x440")
    popup.grab_set()

    ttk.Label(popup, text="Uložené konfigurace operací:", font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, padx=15, pady=10)

    canvas = tk.Canvas(popup, borderwidth=0, highlightthickness=0)
    scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
    scroll_frame = ttk.Frame(canvas)
    scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    pop_win_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(pop_win_id, width=e.width))
    canvas.configure(yscrollcommand=scrollbar.set)
    nastavit_scrolovani(canvas)

    canvas.pack(side="top", fill="both", expand=True, padx=10, pady=5)
    scrollbar.pack(side="right", fill="y")

    def vykreslit():
        for child in scroll_frame.winfo_children():
            child.destroy()
        aktualni = nacist_profily()
        if not aktualni:
            ttk.Label(scroll_frame, text="Žádné profily.").pack(pady=20)
            return

        for p_nazev, p_data in aktualni.items():
            box = ttk.LabelFrame(scroll_frame, text=f" {p_nazev} ", padding=10)
            box.pack(fill=tk.X, expand=True, pady=6, padx=5)
            popis = p_data.get('popis') or "Bez popisu"
            pouzite_s = ", ".join(p_data.get("pouzite_soubory", []))

            ttk.Label(
                box, text=f"Popis: {popis}\nOčekávané soubory: {pouzite_s}",
                font=("Segoe UI", 9, "italic"), foreground="#444444", wraplength=480, justify="left"
            ).pack(anchor=tk.W, pady=(0, 8))

            btns = ttk.Frame(box)
            btns.pack(fill=tk.X)
            ttk.Button(btns, text="Načíst do programu", command=lambda n=p_nazev, d=p_data: aplikovat(n, d, popup)).pack(side=tk.LEFT, padx=(0, 8))
            ttk.Button(btns, text="✏ Upravit", command=lambda n=p_nazev, d=p_data: upravit(n, d, vykreslit)).pack(side=tk.LEFT, padx=4)
            ttk.Button(btns, text="✕ Smazat", command=lambda n=p_nazev: smazat(n, vykreslit)).pack(side=tk.RIGHT, padx=4)

    def aplikovat(nazev, data, win):
        for radek in list(seznam_operaci):
            smazat_radek_operace(radek["frame"], radek)
        for krok in data.get("kroky", []):
            radek = pridat_radek_operace()
            radek["operace"].set(krok.get("operace"))
            zmena_operace(radek["operace"], radek["btn_upravit"], radek)

            for key in ["src_file", "dst_file", "target_file", "klic_src", "klic_dst", "k1_src", "k1_dst", "k2_src", "k2_dst",
                        "konec_col", "mapovani_rules", "klic1", "klic2", "vybrany_sloupec", "hodnota_naplneni", "vlozeny_text",
                        "pn_src", "pn_dst", "rec_col", "filter_text", "sum1", "sum2", "cat_col", "val_col",
                        "m_cat", "m_pct", "e1_cat", "e1_pct", "e2_cat", "e2_pct", "new_col_name", "sum_cols"]:
                if key in krok:
                    radek[key] = krok[key]

            op = krok.get("operace")
            if op == "Kopírovat sloupce (1 kód)":
                radek["btn_upravit"].config(text=f"⚙ Mapování [{radek.get('src_file')}] ➔ [{radek.get('dst_file')}]")
            elif op == "Rozdělit data na IAM / OE / OES":
                radek["btn_upravit"].config(text=f"⚙ [Rozdělit %] {radek.get('src_file')} ➔ {radek.get('dst_file')}")
            elif op == "Sečíst sloupce do nového":
                radek["btn_upravit"].config(text=f"⚙ [{radek.get('target_file')}] Sečíst ➔ [{radek.get('new_col_name')}]")
            elif op == "Kopírovat sloupce (2 kódy)":
                radek["btn_upravit"].config(text=f"⚙ [2 Klíče] [{radek.get('src_file')}] ➔ [{radek.get('dst_file')}]")
            elif op == "Spojit text ze 2 sloupců":
                radek["btn_upravit"].config(text=f"⚙ [{radek.get('target_file')}] Spojit text [{radek.get('klic1')}] + [{radek.get('klic2')}]")

        messagebox.showinfo("Profil načten", f"Profil '{nazev}' byl načten!")
        win.destroy()

    def upravit(stary_nazev, data, refresh_func):
        u_popup = tk.Toplevel(popup)
        u_popup.title(f"Úprava profilu: {stary_nazev}")
        u_popup.geometry("450x220")
        u_popup.grab_set()

        uf = ttk.Frame(u_popup, padding=12)
        uf.pack(fill=tk.BOTH, expand=True)
        ttk.Label(uf, text="Název profilu:", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W)
        u_nazev = ttk.Entry(uf, width=48)
        u_nazev.insert(0, stary_nazev)
        u_nazev.pack(anchor=tk.W, pady=(0, 10))

        ttk.Label(uf, text="Popis:", font=("Segoe UI", 9, "bold")).pack(anchor=tk.W)
        u_popis = ttk.Entry(uf, width=48)
        u_popis.insert(0, data.get("popis", ""))
        u_popis.pack(anchor=tk.W, pady=(0, 10))

        def ulozit_zmeny():
            novy = u_nazev.get().strip()
            if not novy:
                return
            all_p = nacist_profily()
            if stary_nazev in all_p:
                del all_p[stary_nazev]
            all_p[novy] = {"popis": u_popis.get().strip(), "pouzite_soubory": data.get("pouzite_soubory", []), "kroky": data.get("kroky", [])}
            ulozit_profily(all_p)
            u_popup.destroy()
            refresh_func()

        ttk.Button(uf, text="Uložit změny", command=ulozit_zmeny).pack(side=tk.BOTTOM, fill=tk.X, ipady=4)

    def smazat(nazev, refresh_func):
        if messagebox.askyesno("Smazat", f"Opravdu chceš smazat profil '{nazev}'?"):
            all_p = nacist_profily()
            if nazev in all_p:
                del all_p[nazev]
                ulozit_profily(all_p)
                refresh_func()

    vykreslit()


# ================= STRÁNKA 2: PRACOVNÍ PLOCHA =================
def preskladat_radky_operaci(event=None, meneny_radek=None):
    if meneny_radek:
        try:
            nova_pozice = int(meneny_radek["poradi"].get())
            seznam_operaci.remove(meneny_radek)
            seznam_operaci.insert(nova_pozice - 1, meneny_radek)
        except ValueError:
            pass

    for index, radek in enumerate(seznam_operaci):
        radek["poradi"].set(str(index + 1))
        radek["frame"].config(text=f" Krok {index + 1} ")
        radek["frame"].pack_forget()
        radek["frame"].pack(fill=tk.X, expand=True, pady=4, padx=5)

    canvas_kroky.update_idletasks()
    canvas_kroky.configure(scrollregion=canvas_kroky.bbox("all"))


def zmena_operace(cb_operace, btn_upravit, data_radku):
    op = cb_operace.get()
    if op == "Kopírovat sloupce (1 kód)":
        btn_upravit.config(text="⚙ Mapovat sloupce", command=lambda: popup_kopirovat_1klic(data_radku))
    else:
        btn_upravit.config(text="⚙ Nastavit", command=lambda: otevrit_nastaveni_kroku(data_radku))


def pridat_radek_operace():
    radek_frame = ttk.LabelFrame(scrollable_frame_kroky, text=f" Krok {len(seznam_operaci) + 1} ", padding=8)
    radek_frame.pack(fill=tk.X, expand=True, pady=4, padx=5)

    cb_poradi = ttk.Combobox(radek_frame, values=[str(i) for i in range(1, 31)], width=3, state="readonly")
    cb_poradi.set(str(len(seznam_operaci) + 1))
    cb_poradi.pack(side=tk.LEFT, padx=(5, 10))
    cb_poradi.bind("<<ComboboxSelected>>", lambda e: preskladat_radky_operaci(e, data_radku))

    dostupne_operace = [
        "Kopírovat sloupce (1 kód)",
        "Rozdělit data na IAM / OE / OES",
        "Sečíst sloupce do nového",
        "Kopírovat sloupce (2 kódy)",
        "Spojit text ze 2 sloupců",
        "Sloučit duplicitní řádky",
        "Přičíst sloupec k jinému",
        "Smazat obsah sloupce",
        "Vložit stejný text do sloupce",
    ]
    cb_operace = ttk.Combobox(radek_frame, values=dostupne_operace, width=38, state="readonly")
    cb_operace.set("Kopírovat sloupce (1 kód)")
    cb_operace.pack(side=tk.LEFT, padx=5)

    btn_upravit = ttk.Button(radek_frame, text="⚙ Mapovat sloupce")
    btn_upravit.pack(side=tk.LEFT, padx=15)

    prvni_alias = list(nactene_soubory.keys())[0] if nactene_soubory else ""

    data_radku = {
        "frame": radek_frame, "poradi": cb_poradi, "operace": cb_operace, "btn_upravit": btn_upravit,
        "src_file": prvni_alias, "dst_file": prvni_alias, "target_file": prvni_alias,
        "mapovani_rules": [], "klic_src": "", "klic_dst": "", "konec_col": "",
        "klic1": "", "klic2": "", "vybrany_sloupec": "", "hodnota_naplneni": "", "vlozeny_text": "kategorie"
    }

    zmena_operace(cb_operace, btn_upravit, data_radku)
    cb_operace.bind("<<ComboboxSelected>>", lambda e: zmena_operace(cb_operace, btn_upravit, data_radku))

    btn_smazat = ttk.Button(radek_frame, text="✕", width=3, command=lambda: smazat_radek_operace(radek_frame, data_radku))
    btn_smazat.pack(side=tk.RIGHT, padx=5)

    seznam_operaci.append(data_radku)
    preskladat_radky_operaci()
    return data_radku


def smazat_radek_operace(frame, data_radku):
    frame.destroy()
    seznam_operaci.remove(data_radku)
    preskladat_radky_operaci()


# ================= EXECUTION ENGINE (PANDAS / OPENPYXL) =================
def spustit_zpracovani():
    try:
        dfs = {}
        for alias, info in nactene_soubory.items():
            df = nacist_excel(info["cesta"], sheet_name=info["sheet"], header=info["hdr"])
            if df is None:
                messagebox.showerror("Chyba", f"Nepodařilo se načíst soubor '{alias}'.")
                return
            dfs[alias] = df

        for radek in seznam_operaci:
            op = radek["operace"].get()

            # 1. Sloučit duplicity
            if op == "Sloučit duplicitní řádky":
                t_file, odkud = radek.get("target_file"), radek.get("klic1")
                if t_file in dfs and odkud in dfs[t_file].columns:
                    agg_dict = {
                        col: ("sum" if pd.api.types.is_numeric_dtype(dfs[t_file][col]) else "first")
                        for col in dfs[t_file].columns if col != odkud
                    }
                    dfs[t_file] = dfs[t_file].groupby(odkud, as_index=False).agg(agg_dict)

            # 2. Přičíst sloupec k jinému
            elif op == "Přičíst sloupec k jinému":
                t_file, odkud, kam = radek.get("target_file"), radek.get("klic1"), radek.get("klic2")
                if t_file in dfs and odkud in dfs[t_file].columns and kam in dfs[t_file].columns:
                    zdroj = pd.to_numeric(dfs[t_file][odkud], errors="coerce").fillna(0)
                    cil = pd.to_numeric(dfs[t_file][kam], errors="coerce").fillna(0)
                    dfs[t_file][kam] = cil + zdroj

            # 3. Spojit text
            elif op == "Spojit text ze 2 sloupců":
                t_file, cil_col, src_col = radek.get("target_file"), radek.get("klic1"), radek.get("klic2")
                prefix = radek.get("vlozeny_text", "")
                if t_file in dfs and cil_col in dfs[t_file].columns and src_col in dfs[t_file].columns:
                    orig = dfs[t_file][cil_col].fillna("").astype(str)
                    append = dfs[t_file][src_col].fillna("").astype(str)
                    dfs[t_file][cil_col] = orig + " " + prefix + ": " + append

            # 4. Sečíst sloupce do nového
            elif op == "Sečíst sloupce do nového":
                t_file, new_col, cols = radek.get("target_file"), radek.get("new_col_name"), radek.get("sum_cols", [])
                if t_file in dfs and new_col and cols:
                    dfs[t_file][new_col] = dfs[t_file][cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)

            # 5. Kopírovat sloupce (1 kód)
            elif op == "Kopírovat sloupce (1 kód)":
                s_file, d_file = radek.get("src_file"), radek.get("dst_file")
                klic_s, klic_d, konec_col = radek.get("klic_src"), radek.get("klic_dst"), radek.get("konec_col")
                pravidla = radek.get("mapovani_rules", [])

                if s_file in dfs and d_file in dfs and klic_s and klic_d and konec_col:
                    df_src, df_dst = dfs[s_file], dfs[d_file]
                    posledni_idx = df_src[konec_col].dropna().index.max()
                    if pd.isna(posledni_idx):
                        posledni_idx = len(df_src) - 1
                    df_src_oriznuto = df_src.loc[0:posledni_idx]

                    for rule in pravidla:
                        src_col, dst_col = rule["src"], rule["dst"]
                        if src_col in df_src_oriznuto.columns and dst_col in df_dst.columns:
                            mapa = dict(zip(df_src_oriznuto[klic_s], df_src_oriznuto[src_col]))
                            df_dst[dst_col] = df_dst[klic_d].map(mapa).fillna(df_dst[dst_col])

            # 6. Kopírovat sloupce (2 kódy)
            elif op == "Kopírovat sloupce (2 kódy)":
                s_file, d_file = radek.get("src_file"), radek.get("dst_file")
                k1_s, k1_d = radek.get("k1_src"), radek.get("k1_dst")
                k2_s, k2_d = radek.get("k2_src"), radek.get("k2_dst")
                konec_col, pravidla = radek.get("konec_col"), radek.get("mapovani_rules", [])

                if s_file in dfs and d_file in dfs and k1_s and k1_d and k2_s and k2_d:
                    df_src, df_dst = dfs[s_file], dfs[d_file]
                    if konec_col and konec_col in df_src.columns:
                        posledni_idx = df_src[konec_col].dropna().index.max()
                        if pd.isna(posledni_idx):
                            posledni_idx = len(df_src) - 1
                        df_src_oriznuto = df_src.loc[0:posledni_idx]
                    else:
                        df_src_oriznuto = df_src

                    comp_src = df_src_oriznuto[k1_s].astype(str).str.strip().str.lower() + "|||" + df_src_oriznuto[k2_s].astype(str).str.strip().str.lower()
                    comp_dst = df_dst[k1_d].astype(str).str.strip().str.lower() + "|||" + df_dst[k2_d].astype(str).str.strip().str.lower()

                    for rule in pravidla:
                        src_col, dst_col = rule["src"], rule["dst"]
                        if src_col in df_src_oriznuto.columns and dst_col in df_dst.columns:
                            mapa = dict(zip(comp_src, df_src_oriznuto[src_col]))
                            df_dst[dst_col] = comp_dst.map(mapa).fillna(df_dst[dst_col])

            # 7. Rozdělit data na IAM / OE / OES
            elif op == "Rozdělit data na IAM / OE / OES":
                s_file, d_file = radek.get("src_file"), radek.get("dst_file")
                pn_s, pn_d, rec_col = radek.get("pn_src"), radek.get("pn_dst"), radek.get("rec_col")
                f_text = radek.get("filter_text", "").strip()
                sum1, sum2 = radek.get("sum1"), radek.get("sum2")
                cat_col, val_col = radek.get("cat_col"), radek.get("val_col")

                m_cat = radek.get("m_cat", "IAM").strip().upper()
                m_pct = float(radek.get("m_pct", "100") or 0) / 100.0
                e1_cat, e1_pct = radek.get("e1_cat", "OE").strip().upper(), float(radek.get("e1_pct", "30") or 0) / 100.0
                e2_cat, e2_pct = radek.get("e2_cat", "OES").strip().upper(), float(radek.get("e2_pct", "70") or 0) / 100.0

                if s_file in dfs and d_file in dfs and pn_s and pn_d and cat_col and val_col:
                    df_src, df_dst = dfs[s_file], dfs[d_file]
                    dst_pn_clean = df_dst[pn_d].astype(str).str.strip().str.lower()
                    dst_cat_clean = df_dst[cat_col].astype(str).str.strip().str.upper()

                    for idx, row in df_src.iterrows():
                        pn_val = str(row[pn_s]).strip().lower() if pd.notna(row[pn_s]) else ""
                        if not pn_val or pn_val == "nan":
                            continue

                        rec_val = str(row[rec_col]) if rec_col and pd.notna(row[rec_col]) else ""
                        val_sum = 0
                        if sum1 and sum1 in df_src.columns:
                            val_sum += float(pd.to_numeric(row[sum1], errors="coerce") or 0)
                        if sum2 and sum2 != "-- Žádný --" and sum2 in df_src.columns:
                            val_sum += float(pd.to_numeric(row[sum2], errors="coerce") or 0)

                        if f_text and f_text in rec_val:
                            mask = (dst_pn_clean == pn_val) & (dst_cat_clean == m_cat)
                            existing = pd.to_numeric(df_dst.loc[mask, val_col], errors="coerce").fillna(0)
                            df_dst.loc[mask, val_col] = existing + (val_sum * m_pct)
                        else:
                            if e1_cat:
                                mask1 = (dst_pn_clean == pn_val) & (dst_cat_clean == e1_cat)
                                existing1 = pd.to_numeric(df_dst.loc[mask1, val_col], errors="coerce").fillna(0)
                                df_dst.loc[mask1, val_col] = existing1 + (val_sum * e1_pct)
                            if e2_cat:
                                mask2 = (dst_pn_clean == pn_val) & (dst_cat_clean == e2_cat)
                                existing2 = pd.to_numeric(df_dst.loc[mask2, val_col], errors="coerce").fillna(0)
                                df_dst.loc[mask2, val_col] = existing2 + (val_sum * e2_pct)

            # 8. Smazat obsah sloupce
            elif op == "Smazat obsah sloupce":
                t_file, col = radek.get("target_file"), radek.get("vybrany_sloupec")
                if t_file in dfs and col in dfs[t_file].columns:
                    dfs[t_file][col] = None

            # 9. Vložit stejný text do sloupce
            elif op == "Vložit stejný text do sloupce":
                t_file, col = radek.get("target_file"), radek.get("vybrany_sloupec")
                val, konec_c = radek.get("hodnota_naplneni", ""), radek.get("konec_col", "")

                if t_file in dfs and col in dfs[t_file].columns:
                    if konec_c and konec_c != "-- Celý sloupec --" and konec_c in dfs[t_file].columns:
                        posledni_idx = dfs[t_file][konec_c].dropna().index.max()
                        if pd.notna(posledni_idx):
                            dfs[t_file].loc[0:posledni_idx, col] = val
                        else:
                            dfs[t_file][col] = val
                    else:
                        dfs[t_file][col] = val

        # Uložení výsledného Excelu
        if len(nactene_soubory) == 1:
            alias = list(nactene_soubory.keys())[0]
            orig_cesta = nactene_soubory[alias]["cesta"]
            if var_single_output.get() == "overwrite":
                vystupni_cesta = orig_cesta
            else:
                vystupni_cesta = filedialog.asksaveasfilename(
                    filetypes=[("Excel (.xlsx)", "*.xlsx")], defaultextension=".xlsx",
                    initialdir=os.path.dirname(orig_cesta)
                )
                if not vystupni_cesta:
                    return

            ulozit_excel_s_formatem(orig_cesta, vystupni_cesta, dfs[alias], nactene_soubory[alias]["sheet"], nactene_soubory[alias]["hdr"])
        else:
            vystupni_cesta = os.path.splitext(cesta_cil_global)[0] + ".xlsx"
            alias_out = list(nactene_soubory.keys())[-1]
            orig_cesta = nactene_soubory[alias_out]["cesta"]
            ulozit_excel_s_formatem(orig_cesta, vystupni_cesta, dfs[alias_out], nactene_soubory[alias_out]["sheet"], nactene_soubory[alias_out]["hdr"])

        messagebox.showinfo("Hotovo", f"Zpracování proběhlo úspěšně!\nFormátování bylo zachováno.\nUloženo do:\n{os.path.basename(vystupni_cesta)}")
        root.destroy()

    except PermissionError:
        messagebox.showerror("Soubor je otevřený", "Nepodařilo se uložit výsledek!\nCílový soubor je právě otevřený v jiném programu (např. v Excelu).\nNejdříve ho prosím zavři.")
    except Exception as e:
        messagebox.showerror("Chyba při zpracování", f"Něco kleklo v motoru automatizace:\n{str(e)}")


def navrat_na_krok_1():
    frame_strana2.pack_forget()
    frame_strana1.pack(fill=tk.BOTH, expand=True)


# ================= HLAVNÍ OKNO PROGRAMU =================
root = tk.Tk()
root.title("Automatizace konverze tabulek")
root.geometry("960x650")

style = ttk.Style()
style.theme_use("vista")

# --- STRÁNKA 1: SPRÁVA SOUBORŮ ---
frame_strana1 = ttk.Frame(root, padding="20")
frame_strana1.pack(fill=tk.BOTH, expand=True)
ttk.Label(frame_strana1, text="Krok 1: Správa vstupních souborů", font=("Segoe UI", 12, "bold")).pack(anchor=tk.W, pady=(0, 10))

canvas_soubory = tk.Canvas(frame_strana1, borderwidth=0, highlightthickness=0)
scrollbar_soubory = ttk.Scrollbar(frame_strana1, orient="vertical", command=canvas_soubory.yview)
scroll_frame_soubory = ttk.Frame(canvas_soubory)

scroll_frame_soubory.bind("<Configure>", lambda e: canvas_soubory.configure(scrollregion=canvas_soubory.bbox("all")))
win_soubory_id = canvas_soubory.create_window((0, 0), window=scroll_frame_soubory, anchor="nw")
canvas_soubory.bind("<Configure>", lambda e: canvas_soubory.itemconfig(win_soubory_id, width=e.width))

canvas_soubory.configure(yscrollcommand=scrollbar_soubory.set)
nastavit_scrolovani(canvas_soubory)

canvas_soubory.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, 10))
scrollbar_soubory.pack(side=tk.RIGHT, fill=tk.Y)

ttk.Button(frame_strana1, text="+ Přidat soubor", command=pridat_soubor_ui).pack(anchor=tk.W, pady=(0, 10))

f_vystup_container = ttk.Frame(frame_strana1)
f_vystup_container.pack(fill=tk.X, pady=(0, 15))

box_vystup_single = ttk.LabelFrame(f_vystup_container, text=" Volba uložení pro 1 soubor ", padding="10")
var_single_output = tk.StringVar(value="copy")
ttk.Radiobutton(box_vystup_single, text="Uložit jako nový soubor (Kopie)", variable=var_single_output, value="copy").pack(anchor=tk.W, pady=2)
ttk.Radiobutton(box_vystup_single, text="Přepsat původní originální soubor", variable=var_single_output, value="overwrite").pack(anchor=tk.W, pady=2)

box_vystup_multi = ttk.LabelFrame(f_vystup_container, text=" Výstupní soubor ", padding="10")
ttk.Button(box_vystup_multi, text="Určit kam uložit nový soubor...", command=vybrat_cilovy_soubor).pack(side=tk.LEFT, padx=(0, 10))
lbl_cil_path = ttk.Label(box_vystup_multi, text="Není vybráno", font=("Segoe UI", 9, "italic"), foreground="gray")
lbl_cil_path.pack(side=tk.LEFT)

pridat_soubor_ui()

ttk.Button(frame_strana1, text="Pokračovat k nastavení automatizace ➔", command=prejit_na_pracovni_plochu).pack(fill=tk.X, side=tk.BOTTOM, ipady=8)


# --- STRÁNKA 2: PRACOVNÍ PLOCHA KROKŮ ---
frame_strana2 = ttk.Frame(root, padding="15")
frame_body = ttk.Frame(frame_strana2)
frame_body.pack(fill=tk.BOTH, expand=True)

frame_left = ttk.Frame(frame_body)
frame_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
ttk.Label(frame_left, text="Seznam prováděných kroků automatizace", font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=(0, 5))

canvas_kroky = tk.Canvas(frame_left, borderwidth=0, highlightthickness=0)
scrollbar_kroky = ttk.Scrollbar(frame_left, orient="vertical", command=canvas_kroky.yview)
scrollable_frame_kroky = ttk.Frame(canvas_kroky)

scrollable_frame_kroky.bind("<Configure>", lambda e: canvas_kroky.configure(scrollregion=canvas_kroky.bbox("all")))
main_canvas_win = canvas_kroky.create_window((0, 0), window=scrollable_frame_kroky, anchor="nw")
canvas_kroky.bind("<Configure>", lambda e: canvas_kroky.itemconfig(main_canvas_win, width=e.width))

canvas_kroky.configure(yscrollcommand=scrollbar_kroky.set)
nastavit_scrolovani(canvas_kroky)

canvas_kroky.pack(side="left", fill="both", expand=True)
scrollbar_kroky.pack(side="right", fill="y")

# Pravý panel
frame_right = ttk.Frame(frame_body, width=200)
frame_right.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))

box_profily = ttk.LabelFrame(frame_right, text=" Profily ", padding="10")
box_profily.pack(fill=tk.X, pady=(0, 15))
ttk.Button(box_profily, text="💾 Uložit profil", command=popup_ulozit_profil).pack(fill=tk.X, pady=3, ipady=3)
ttk.Button(box_profily, text="📂 Načíst profil", command=popup_nacist_profil).pack(fill=tk.X, pady=3, ipady=3)

box_kroky = ttk.LabelFrame(frame_right, text=" Úprava kroků ", padding="10")
box_kroky.pack(fill=tk.X, pady=(0, 15))
ttk.Button(box_kroky, text="+ Přidat další krok", command=pridat_radek_operace).pack(fill=tk.X, pady=3, ipady=3)

box_akce = ttk.LabelFrame(frame_right, text=" Akce ", padding="10")
box_akce.pack(fill=tk.X, side=tk.BOTTOM)
ttk.Button(box_akce, text="🚀 Spustit konverzi", command=spustit_zpracovani).pack(fill=tk.X, pady=(3, 8), ipady=6)
ttk.Button(box_akce, text="⮌ Zpět na výběr souborů", command=navrat_na_krok_1).pack(fill=tk.X, pady=3, ipady=3)

root.mainloop()